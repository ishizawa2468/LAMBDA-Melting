"""
nxsとponiを読み込んで，積算・cakingされたデータを返す
"""
import os
import openpyxl
import numpy as np
import pandas as pd
from tqdm import tqdm

import h5py
from pyFAI.azimuthalIntegrator import AzimuthalIntegrator
from modules.file_format.HDF5 import HDF5Reader


class NxsFile(HDF5Reader):
    def __init__(self, nxs_path=None, poni_path=None, mask_path=None, run_name=None):
        super().__init__(file_path=nxs_path) # HDF5Readerの機能を持つ
        self.data_path_to_detector = "/entry/instrument/detector" # hdfとしてのデータまでのpath
        self.run_name = run_name # これはponiを探す時に使う

        # nxsをもとに実行
        self._read_params()
        self._create_integrator() # AzimuthalIntegratorを作成する

        # 分割数の設定
        self.npt_rad = 5_000
        self.npt_azim = 5_000

        # poniとmaskの設定(オプション)
        if poni_path is not None:
            self.set_poni(poni_path=poni_path)
        if mask_path is not None:
            self.set_mask(mask_path=mask_path)


    def _read_params(self):
        with h5py.File(self.file_path, 'r') as f:
            self.frame_num = f[os.path.join(self.data_path_to_detector, 'data')].shape[0]
            self.exposure_ms = f.get(os.path.join(self.data_path_to_detector, 'count_time'))[0]
        self.fps = 1_000.0 / self.exposure_ms

    def _create_integrator(self):
        ai = AzimuthalIntegrator()
        self.ai = ai

    # def _get_run_date(self):
    #     # excelからponi_pathを取得する
    #     # excelファイルへのpath
    #     excel_path = os.path.join(self.setting.PATH_TO_RUN_PARAMS, "run_list.xlsx")
    #     # excel sheetから(year, month)に一致する行を見つける
    #     workbook = openpyxl.load_workbook(excel_path)
    #     sheet = workbook['run_year_month']
    #     run_row = None # これに目当ての行番号を入れる
    #     for row in range(1, sheet.max_row + 1):
    #         if sheet[f'C{row}'].value == self.run_name:
    #             run_row = row+1
    #             break
    #     # 見つからなかったらエラーを出す
    #     if run_row is None:
    #         raise ValueError(f"{self.run_name} が run_list.xlsx の run_year_monthシートで見つかりませんでした。")
    #     # 見つかったらyear, monthを読み込む
    #     self.year = sheet[f"D{run_row}"].value
    #     self.month = sheet[f"E{run_row}"].value

    # def _get_poni_name(self):
    #     # excelからponi_pathを取得する
    #     # excelファイルへのpath
    #     excel_path = os.path.join(self.setting.PATH_TO_RUN_PARAMS, "run_list.xlsx")
    #     # excel sheetから(year, month)に一致する行を見つける
    #     workbook = openpyxl.load_workbook(excel_path)
    #     sheet = workbook['calib_relation']
    #     run_row = None # これに目当ての行番号を入れる
    #     for row in range(1, sheet.max_row + 1):
    #         if (sheet[f'A{row}'].value == self.year) and (sheet[f'B{row}'].value == self.month):
    #             run_row = row
    #             break
    #     # 見つからなかったらエラーを出す
    #     if run_row is None:
    #         raise ValueError(f"{self.run_name} の月日 ({self.year}, {self.month}) が run_list.xlsx の calib_relationシートで見つかりませんでした。")
    #     # 見つかったらponi名を読み込む
    #     poni_name = sheet[f'C{run_row}'].value
    #     return poni_name
    #
    # def _find_mask(self):
    #     return os.path.join(self.setting.PATH_TO_REF_DATA, 'lambda_mask.npy')

    """maskファイルの設定"""
    def set_mask(self, *, mask_path=None):
        if mask_path.endswith('.npy'):
            print(f" > Set mask: {mask_path}")
            self.ai.mask = np.load(mask_path) # AzimuthalIntegratorを作成した後でないとエラー
        else:
            raise Exception(f'Mask path {mask_path} not .npy')

    """poniファイルの設定"""
    def set_poni(self, *, poni_path=None):
        # FIXME エラー処理不要
        if poni_path is None:
            raise ValueError("poni_pathがNoneです")
        else:
            self.poni_path = poni_path
            print(f" > Set poni: {self.poni_path}")
            self.ai.load(self.poni_path)

    def get_tth(self, npt_rad=None):
        # 分割数を指定されていなければ、クラスに設定された分割数を用いる
        npt_rad = npt_rad if npt_rad is not None else self.npt_rad
        try:
            frame_data = self._read_frame_data(0) # 0frame目のデータを読み込む
            tth, I = self.ai.integrate1d(frame_data, npt=npt_rad, unit="2th_deg")
            return tth
        except Exception as e:
            raise RuntimeError(f"Frame 0 の 1D 積分中にエラーが発生しました: {str(e)}")

    def get_azi(self, npt_rad=None, npt_azim=None):
        # 分割数のデフォルト設定
        npt_rad = npt_rad if npt_rad is not None else self.npt_rad
        npt_azim = npt_azim if npt_azim is not None else self.npt_azim

        try:
            frame_data = self._read_frame_data(0) # 0frame目のデータを読み込む
            I, tth, azi = self.ai.integrate2d(frame_data,
                                              npt_rad=npt_rad,  # NOTE これはintegrate_1dと揃える
                                              npt_azim=npt_azim,
                                              unit="2th_deg")
            return azi
        except Exception as e:
            raise RuntimeError(f"Frame 0 の 2D 積分中にエラーが発生しました: {str(e)}")

    def get_1d_pattern_data(self, frame=None, npt_rad=None):
        """
        指定したフレームのデータを1次元に積分するメソッド

        Parameters:
        frame (int): 積分するフレームのインデックス
        npt_rad (int): 回折角方向の分割数

        Returns:
            : 強度
        """
        # 分割数を指定されていなければ、クラスに設定された分割数を用いる
        npt_rad = npt_rad if npt_rad is not None else self.npt_rad

        # 読み込み
        try:
            frame_data = self._read_frame_data(frame)
            tth, I = self.ai.integrate1d(frame_data, npt=npt_rad, unit="2th_deg")
            return I # tthは別でメソッドを作っている
        except Exception as e:
            raise RuntimeError(f"Frame {frame} の 1D 積分中にエラーが発生しました: {str(e)}")

    def get_caked_data(self, frame, npt_rad=None, npt_azim=None):
        """
        指定したフレームのデータを2次元に積分し、cakingを行うメソッド

        Parameters:
        frame (int): 積分するフレームのインデックス
        npt_rad (int): 回折角方向の分割数（デフォルトはself.npt_rad）
        npt_azim (int): 方位角方向の分割数（デフォルトはself.npt_azim）

        Returns:
            : 強度
        """
        # 分割数のデフォルト設定
        npt_rad = npt_rad if npt_rad is not None else self.npt_rad
        npt_azim = npt_azim if npt_azim is not None else self.npt_azim

        # 読み込み
        try:
            frame_data = self._read_frame_data(frame)
            # 2D積分を行いcakedデータを取得する
            I, tth, azi = self.ai.integrate2d(frame_data,
                                              npt_rad=npt_rad,  # NOTE これはintegrate_1dと揃える
                                              npt_azim=npt_azim,
                                              unit="2th_deg")
            return I
        except Exception as e:
            raise RuntimeError(f"Frame {frame} の 2D 積分中にエラーが発生しました: {str(e)}")

    def _read_frame_data(self, frame):
        # 複数の露光データがあるとき、最初のframeを飛ばす。使い物にならないときがある&重要でないことが多いため。
        if frame == 0 and self.frame_num > 1:
            frame = 1

        with h5py.File(self.file_path, 'r') as f:
            frame_data = f[os.path.join(self.data_path_to_detector, 'data')][frame, :, :]
        return frame_data
